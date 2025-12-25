from openpyxl import load_workbook
from pathlib import Path


class ExcelWriter:
    """
    Responsible ONLY for writing results into results.xlsx
    Does NOT calculate or read configs.
    """

    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)

        if not self.excel_path.exists():
            raise FileNotFoundError(
                f"Results sheet not found: {self.excel_path}"
            )

        self.wb = load_workbook(self.excel_path)
        self.ws = self.wb.active


    # --------------------------
    # HELPERS
    # --------------------------

    def _find_config_row(self, config_id: str) -> int:
        """
        Find row where config_id exists in column A
        """

        for row in range(3, self.ws.max_row + 1):
            val = self.ws.cell(row=row, column=1).value
            if val == config_id:
                return row

        raise ValueError(f"Config ID '{config_id}' not found in sheet")


    def _find_audio_block_start(self, audio_id: str) -> int:
        """
        Returns starting column index of audio block (wer/der/rtf)
        Row2 contains repeating headers
        """

        for col in range(5, self.ws.max_column + 1):
            header = self.ws.cell(row=1, column=col).value
            if header == audio_id:
                return col

        raise ValueError(f"Audio '{audio_id}' not found in Excel header")


    # --------------------------
    # PUBLIC API
    # --------------------------

    def write_audio_result(self,
                        config_id: str,
                        audio_id: str,
                        wer: float,
                        der: float,
                        rtf: float):
        """
        Writes metrics in:
        row   = config_id
        block = audio_id (WER / DER / RTF)
        """

        row = self._find_config_row(config_id)
        col = self._find_audio_block_start(audio_id)

        self.ws.cell(row=row, column=col + 0).value = wer
        self.ws.cell(row=row, column=col + 1).value = der
        self.ws.cell(row=row, column=col + 2).value = rtf

        self.wb.save(self.excel_path)


    def write_overall_result(self,
                            config_id: str,
                            wer: float,
                            der: float,
                            rtf: float):
        """
        Writes overall dataset scores in columns:
        B=WER  C=DER  D=RTF
        """

        row = self._find_config_row(config_id)

        self.ws.cell(row=row, column=2).value = wer
        self.ws.cell(row=row, column=3).value = der
        self.ws.cell(row=row, column=4).value = rtf

        self.wb.save(self.excel_path)
